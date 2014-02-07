#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
from openpyxl import load_workbook

def get_headings(workbook):
    col_index = 0
    headings = []
    while True:
        heading = self.worksheet.cell(row=0, column=col_index).value
        if heading:
            headings.append(heading)
            col_index += 1
        else:
            break
        return headings

wb = load_workbook(os.path.join(os.path.dirname(__file__), '../data/schoenbergAll.xlsx'))

print get_headings(wb)
